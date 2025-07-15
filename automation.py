import openpyxl

# === Description ===
# This script reads data from a central source Excel file and populates specific target cells
# in a series of Sales Order (SO) files, streamlining the manual reconciliation process.

# === Load source workbook ===
source_workbook = openpyxl.load_workbook("path/to/source/SO_Transactions.xlsx")
source_sheet = source_workbook.active

# === Define target files and target cells to update ===
target_files = {
    "path/to/Folder/File_01.xlsx": "I17",
    "path/to/Folder/File_02.xlsx": "I9",
    "path/to/Folder/File_03.xlsx": "I9",
    "path/to/Folder/File_04.xlsx": "I9",
    "path/to/Folder/File_05.xlsx": "I9",
    "path/to/Folder/File_06.xlsx": "I9"
}

# === Define the row and column in the source sheet to pull data from ===
source_row = 2  # example row index
source_col = 7  # column 'G'

# === Loop through each target file and update the specified cell ===
for file_path, cell in target_files.items():
    try:
        target_workbook = openpyxl.load_workbook(file_path)
        target_sheet = target_workbook.active

        value = source_sheet.cell(row=source_row, column=source_col).value
        target_sheet[cell] = value

        target_workbook.save(file_path)
        print(f"Updated {file_path} at {cell} with value: {value}")
    except Exception as e:
        print(f"Failed to update {file_path}: {e}")